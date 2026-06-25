"""Excel export schema per Section 5 of spec (columns A-I)."""
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# BUG 3 FIX: sys.path.insert removed — path manipulation belongs in entrypoints,
# not in library modules that are imported by the application server.

# Column definitions: (header label, column width)
HEADERS = [
    ("Account / Parcel ID",        22),
    ("Site Address",               32),
    ("Distressed Status",          28),
    ("Estimated Repair Cost",      20),
    ("After-Repair Value (ARV)",   22),
    ("Owner Legal Name",           28),
    ("Verified Mobile Line 1",     20),
    ("Verified Mobile Line 2",     20),
    ("Primary Active Email",       32),
]

# Palette — stays consistent with the app's terminal skin
_BLACK      = "0A0A0A"
_DARK_ROW   = "111111"   # zebra stripe dark
_LIGHT_ROW  = "1A1A1A"   # zebra stripe light (subtle — still dark skin)
_LIME       = "DEFF9A"   # accent for header text
_WHITE      = "FFFFFF"
_BORDER_CLR = "2A2A2A"


def build_export(properties: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Off-Market Opportunities"

    # ── Shared style objects ──────────────────────────────────────────────────
    border_side  = Side(border_style="thin", color=_BORDER_CLR)
    cell_border  = Border(
        left=border_side, right=border_side,
        top=border_side,  bottom=border_side,
    )
    center       = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    money_fmt    = '"$"#,##0'

    # ── Row 1: Title ──────────────────────────────────────────────────────────
    # BUG 1 FIX: No blank row 2 — title on row 1, headers on row 2, data from row 3.
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value     = "OFF-MARKET PROPERTY INTELLIGENCE — EXPORT"
    title_cell.font      = Font(bold=True, size=13, name="Calibri", color=_LIME)
    # BUG 2 FIX: Title cell now has a fill so the visual hierarchy makes sense
    title_cell.fill      = PatternFill("solid", fgColor=_BLACK)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Row 2: Column headers ─────────────────────────────────────────────────
    header_font = Font(bold=True, color=_WHITE, size=10, name="Calibri")
    header_fill = PatternFill("solid", fgColor=_BLACK)

    for col_idx, (label, width) in enumerate(HEADERS, start=1):
        cell             = ws.cell(row=2, column=col_idx, value=label)
        cell.font        = header_font
        cell.fill        = header_fill
        cell.alignment   = center
        cell.border      = cell_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[2].height = 34

    # ── Freeze above data rows ────────────────────────────────────────────────
    # BUG 7 NOTE: freeze at A3 — keeps title + header always visible while scrolling
    ws.freeze_panes = "A3"

    # ── Data rows ─────────────────────────────────────────────────────────────
    # BUG 6 FIX: alternating row fill for readability
    fill_dark  = PatternFill("solid", fgColor=_DARK_ROW)
    fill_light = PatternFill("solid", fgColor=_LIGHT_ROW)

    for row_idx, p in enumerate(properties, start=3):
        underwrite = p.get("underwrite") or {}
        repair     = (underwrite.get("repair") or {}).get("total_repair_cost", "")
        arv        = (underwrite.get("arv")    or {}).get("arv", "")

        skip    = p.get("skip_trace_data") or {}
        mobiles = skip.get("mobile_lines", []) or []
        emails  = skip.get("emails", [])       or []

        # BUG 5 FIX: defensive .get("number", "") — no KeyError if key absent
        mobile_1 = mobiles[0].get("number", "") if len(mobiles) > 0 else ""
        mobile_2 = mobiles[1].get("number", "") if len(mobiles) > 1 else ""
        email_1  = emails[0] if emails else ""

        # BUG 4 FIX: zip_code not zip — matches the property data model
        address = (
            f"{p.get('site_address', '')}, "
            f"{p.get('city', '')}, "
            f"{p.get('state', '')} "
            f"{p.get('zip_code', '')}"
        ).strip(", ")

        values = [
            p.get("apn", ""),
            address,
            ", ".join(p.get("distress_statuses") or []) or p.get("primary_status", ""),
            repair,
            arv,
            p.get("owner_name", ""),
            mobile_1,
            mobile_2,
            email_1,
        ]

        row_fill = fill_dark if row_idx % 2 == 0 else fill_light

        for col_idx, val in enumerate(values, start=1):
            cell             = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border      = cell_border
            cell.alignment   = left_align
            cell.fill        = row_fill
            cell.font        = Font(name="Calibri", size=10, color=_WHITE)
            # Apply money format to repair cost (col 4) and ARV (col 5)
            if col_idx in (4, 5) and isinstance(val, (int, float)):
                cell.number_format = money_fmt

    # ── Serialize ─────────────────────────────────────────────────────────────
    # BUG 8 FIX: use context manager so BytesIO is explicitly closed after read
    with BytesIO() as bio:
        wb.save(bio)
        bio.seek(0)
        return bio.read()
